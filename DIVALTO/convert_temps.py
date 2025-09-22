import pandas as pd
from datetime import datetime
import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- Choix du fichier source ---
if len(sys.argv) > 1:
    fichier_source = sys.argv[1]
else:
    fichiers = glob.glob("*.xlsx")
    if not fichiers:
        print("❌ Aucun fichier Excel trouvé dans le dossier.")
        sys.exit(1)
    fichiers.sort(key=os.path.getmtime, reverse=True)
    fichier_source = fichiers[0]

print(f"➡ Traitement du fichier : {fichier_source}")

# --- Lecture ---
df = pd.read_excel(fichier_source)

# --- Conversion de la colonne Temps réalisé ---
def convertir_en_heures(val):
    if pd.isna(val):
        return 0
    try:
        h, m, s = map(int, str(val).split(":"))
        return h + m/60 + s/3600
    except Exception:
        return 0

df["Temps réalisé"] = df["Temps réalisé"].apply(convertir_en_heures)

# --- Nom du fichier de sortie ---
horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
nom, ext = os.path.splitext(os.path.basename(fichier_source))
fichier_sortie = f"{nom}_converti_{horodatage}{ext}"

# --- Sauvegarde temporaire avec pandas ---
df.to_excel(fichier_sortie, index=False)

# --- Post-traitement avec openpyxl ---
wb = load_workbook(fichier_sortie)
ws = wb.active

# Police Arial pour toutes les cellules
for row in ws.iter_rows():
    for cell in row:
        cell.font = Font(name="Arial", size=10)

# Ajustement automatique de la largeur des colonnes
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Sauvegarde finale
wb.save(fichier_sortie)

print(f"✅ Conversion terminée. Fichier créé : {fichier_sortie}")
